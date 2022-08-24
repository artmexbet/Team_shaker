import openpyxl as xl

from styles import *

#  Подгрузка мест
with open("points.csv", encoding="utf8") as f:
    formatted_points = [[k.strip() for k in i.split(",")] for i in f.readlines()[1:]]

teams = ["Альфа", "Бета", "Гамма", "Тета", "Эпсилон", "Дельта", "Йота", "Каппа", "Эта", "Кси", "Омикрон", "Омега"]

wb = xl.Workbook()
for team in teams:
    sheet = wb.create_sheet(team, 0)

    # Меняем ширину столбцов
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 20

    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Команда " + sheet.title
    sheet["A1"].font = BOLD_FONT

    sheet["A2"] = "Место"
    sheet["A2"].font = BOLD_FONT

    sheet["B2"] = "Этап"
    sheet["B2"].font = BOLD_FONT

    sheet["C2"] = "Оценка"
    sheet["C2"].font = BOLD_FONT

    sheet["D2"] = "Подпись"
    sheet["D2"].font = BOLD_FONT

    count = 3  # Стартуем с 3, потому что 1 и 2 строки заняты шапкой
    for place, point in formatted_points:
        sheet[f"A{count}"] = place
        sheet[f"A{count}"].font = FONT

        sheet[f"B{count}"] = point
        sheet[f"B{count}"].font = FONT

        sheet.row_dimensions[count].height = 25

        count += 1

    # Расставляем стили всем задействемым ячейкам
    for row in sheet.rows:
        for col in row:
            col.border = BORDER
            col.alignment = CENTER_ALIGNMENT

    formatted_points.append(formatted_points.pop(0))  # Переставляем первый элемент в конец
wb.save("test.xlsx")
